# author: Chengrui Dai(SA)
import yara
import os
import openpyxl
import sys

class FeatureHelper(object):
    """

    """
    def __init__(self):
        self.root_path_ = os.path.split(os.path.realpath(__file__))[0]
        self.mk = []
        self.hk = []
        self.jk = []
        self.lk = []
        self.vk = []
        self.pk = []
        self.ok = []
        self.mk_start = 1
        self.hk_start = 101
        self.jk_start = 201
        self.lk_start = 301
        self.vk_start = 401
        self.pk_start = 501
        self.ok_start = 601
        self.yara_path = os.path.join(self.root_path_,r"yara_ssa.yar")
        self.wiki_path = os.path.join(self.root_path_,r"wiki_format")
        self.feature_xlsx = os.path.join(self.root_path_,r"feature.xlsx")
        self.clear_env()

    def parse_xlsx(self):
        wb = openpyxl.load_workbook(self.feature_xlsx)  
        sheets = wb.get_sheet_names()  
        for i in range(len(sheets)):
            sheet= wb.get_sheet_by_name(sheets[i])
            if sheet.title == "Malicious Keywords":
                for r in range(1,sheet.max_row+1):
                    self.mk.append(str(sheet.cell(row = r,column = 1).value))
            if sheet.title == "HTML Keywords":
                for r in range(1,sheet.max_row+1):
                    self.hk.append(str(sheet.cell(row = r,column = 1).value))
            if sheet.title == "JS API Function Keywords":
                for r in range(1,sheet.max_row+1):
                    self.jk.append(str(sheet.cell(row = r,column = 1).value))
            if sheet.title == "Local API Function Keywords":
                for r in range(1,sheet.max_row+1):
                    self.lk.append(str(sheet.cell(row = r,column = 1).value))	    
            if sheet.title == "VBS Keywords":
                for r in range(1,sheet.max_row+1):
                    self.vk.append(str(sheet.cell(row = r,column = 1).value))
            if sheet.title == "Powershell Keywords":
                for r in range(1,sheet.max_row+1):
                    self.pk.append(str(sheet.cell(row = r,column = 1).value))
            if sheet.title == "Other Keywords":
                for r in range(1,sheet.max_row+1):
                    self.ok.append(str(sheet.cell(row = r,column = 1).value))

    def clear_env(self):
        if os.path.exists(self.yara_path):
            os.remove(self.yara_path)	   
        if os.path.exists(self.wiki_path):
            os.remove(self.wiki_path) 			

    def generate_yara(self):
        self.make_yara(self.mk_start,self.mk)
        self.make_yara(self.hk_start,self.hk)
        self.make_yara(self.jk_start,self.jk)
        self.make_yara(self.lk_start,self.lk)
        self.make_yara(self.vk_start,self.vk)
        self.make_yara(self.pk_start,self.pk)
        self.make_yara(self.ok_start,self.ok)

    def make_yara(self,start_index,keywords):
        index = start_index
        with open(self.yara_path,'a') as ya:
            for info in keywords:
                rule_info = info.replace(' ','_').replace('.','_')
            
                ya.writelines(\
"rule " + rule_info + '\n\
{\n\
    meta:\n\
        index = ' + str(index) + '\n\
    strings:\n\
        $s1 = "' + info + '"\n\
    condition:\n\
        all of them\n\
}\n'\
                            )
                ya.writelines('\n')
                index += 1        

    def generate_wiki(self):
        self.make_wiki(self.mk_start,self.mk)
        self.make_wiki(self.hk_start,self.hk)
        self.make_wiki(self.jk_start,self.jk)
        self.make_wiki(self.lk_start,self.lk)
        self.make_wiki(self.vk_start,self.vk)
        self.make_wiki(self.pk_start,self.pk)
        self.make_wiki(self.ok_start,self.ok)

    def make_wiki(self,start_index,keywords):
        index = start_index
        with open(self.wiki_path,'a') as wi:
            if index == self.mk_start:
                wi.writelines("Malicious Keywords\n")
            elif index == self.hk_start:
                wi.writelines("HTML Keywords\n")
            elif index == self.jk_start:
                wi.writelines("JavaScript API Function Keywords\n")
            elif index == self.lk_start:
                wi.writelines("Local JavaScript/VBS API Function Keywords\n")
            elif index == self.vk_start:
                wi.writelines("VBScript Keywords\n")
            elif index == self.pk_start:
                wi.writelines("Powershell Keywords\n")
            elif index == self.ok_start:
                wi.writelines("Other Keywords\n")
            for info in keywords:
                wi.writelines('"' + info + '",')
                index += 1
                if (index - 1) % 5 == 0:
                    wi.writelines(" // " + str(index - 6) + "-" + str(index - 2))
                    wi.writelines('\n')
        
            if (index - 1) % 5 != 0:
                if (index - 1) % 5 + 1 != 2:
                    wi.writelines(" // " + str(index - (index - 1) % 5 - 1) + "-" + str(index - 2))
                else:
                    wi.writelines(" // " + str(index - 2))
            wi.writelines('\n\n')

    def process(self,feature_path):
        self.feature_xlsx = feature_path
        self.parse_xlsx()
        self.generate_yara()
        self.generate_wiki()


def main():
    Fh = FeatureHelper()
    Fh.process(sys.argv[1])


if __name__ == '__main__':
    main()
