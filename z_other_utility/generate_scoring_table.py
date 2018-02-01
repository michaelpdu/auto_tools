import os
import openpyxl
import re
import sys
import time

class FillScoringTable:
    def __init__(self):
#        self.srcDir = r'\\nj-daisy-li2\BES\MachineLearning\ML-Decision-Engine\console_output'
        self.tablePath = 'scoring_table_{}.xlsx'.format(time.strftime('%Y%m%d%H%M%S', time.localtime()))
        self.srcDir = None
        self.content_to_find = []
        self.title_to_find = []
        self.testPath = r'\\nj-daisy-li2\BES\MachineLearning\ML-Decision-Engine\console_output\xgboost_stdout.txt'
        self.filename_to_row = {}

    def generateXlsx(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['C1'] = 'Training-Testing Set'
        ws['H1'] = 'Validation Set'
        ws['C2'] = 'Malicious Num.'
        ws['D2'] = 'Normal Num.'
        ws['E2'] = 'Best Params.'
        ws['F2'] = 'Training Time Delta'
        ws['G2'] = 'Model Size'
        ws['H2'] = 'Malicious Num.'
        ws['I2'] = 'Normal Num.'
        ws['J2'] = 'Validation Time Delta'
        ws['K2'] = 'TP'
        ws['L2'] = 'FP'
        ws['M2'] = 'TN'
        ws['N2'] = 'FN'
        ws['O2'] = 'Accuracy'
        ws['P2'] = 'Precision(PPV)'
        ws['Q2'] = 'Recall'
        ws['R2'] = 'FPR(FP1)'
        ws['S2'] = 'FDR(FP2|1-PPV)'
        ws['T2'] = 'F1-Measure'
        ws['A3'] = 'LibSVM'
        ws['A4'] = 'LibSVM'
        ws['A5'] = 'XGBoost'
        ws['A6'] = 'Keras'
        ws['A7'] = 'Keras'
        ws['B3'] = 'Linear Kernel'
        ws['B4'] = 'RBF Kernel'
        ws['B6'] = '5-layers dropout'
        ws['B7'] = '7-layers dropout'
        wb.save(filename=self.tablePath)

    def parserXlsx(self,Path):
        wb = openpyxl.load_workbook(filename=Path)
        sheets = wb.get_sheet_names()
        sheet1 = sheets[0]
        ws = wb.get_sheet_by_name(sheet1)
        row_to_find = 2
        for i in range(1,len(ws[row_to_find])+1):
            value = ws.cell(row=row_to_find,column=i).value
            if value is not None:
                self.content_to_find.append(value.encode('unicode-escape').decode('string_escape').split('.')[0])
#        print self.content_to_find

    def getFileName(self,Dir):
        for file_name_tmp in os.listdir(Dir):
            file_name_split = file_name_tmp.split('_')
            file_name_length = len(file_name_split)
            if file_name_split[file_name_length - 1] == 'stdout.txt':
                file_name = os.path.join(self.srcDir, file_name_tmp)
                if file_name_split[file_name_length - 2] == 'keras':
                    self.filename_to_row[file_name] = 6
                elif file_name_split[file_name_length - 2] == 'xgboost':
                    self.filename_to_row[file_name] = 5
                elif file_name_split[file_name_length - 2] == 'rbf':
                    self.filename_to_row[file_name] = 4
                elif file_name_split[file_name_length - 2] == 'linear':
                    self.filename_to_row[file_name] = 3
                else:
                    pass
#        print self.filename_to_row

    def adpater(self,filePath,content_name):
        value = None
        if content_name == 'Validation Time Delta':
            value = self.findInFile(filePath,'Scoring Time Delta')
        elif content_name == 'Malicious Num':
            value_tmp1 = self.findInFile(filePath, 'TP')
            value_tmp2 = self.findInFile(filePath, 'FN')
            if value_tmp1 is not None and value_tmp2 is not None:
                value_tmp12 = int(value_tmp1) + int(value_tmp2)
                value = str(value_tmp12)
        elif content_name == 'Normal Num':
            value_tmp3 = self.findInFile(filePath, 'TN')
            value_tmp4 = self.findInFile(filePath, 'FP')
            if value_tmp3 is not None and value_tmp4 is not None:
                value_tmp34 = int(value_tmp3) + int(value_tmp4)
                value = str(value_tmp34)
        else:
            value = self.findInFile(filePath, content_name)
        return value

    def findInFile(self,Path,content_to_find_tmp):
        file_to_read = open(Path,'rb+')
        all_line = file_to_read.read()
        p = re.compile('\(')
        content_to_find_tmp1 = p.sub('\\(', content_to_find_tmp)
        p1 = re.compile('\|')
        content_to_find_tmp2 = p1.sub('\\|', content_to_find_tmp1)
        p2 = re.compile('\)')
        content_to_find_tmp3 = p2.sub('\\)', content_to_find_tmp2)
        if content_to_find_tmp3 == 'Best Params':
            content_to_find = content_to_find_tmp3 + ':[ ]?(\{.*\})'
        else:
            content_to_find = content_to_find_tmp3 + ':[ ]?(\d*[.]?\d*)'
        pattern = re.compile(content_to_find,re.IGNORECASE)
        m = pattern.search(all_line)
        if m:
            return m.group(1)
        else:
            return None

help_msg="""
    Usage:
        python generate_scoring_table.py src_dir 
    Note:
        src_dir is a directory, which contains many file:
            src_dir
                |- file_name_1
                |- file_name_2
                |- ...
                |- file_name_m
"""

if __name__ == '__main__':
    fill_scoring_table = FillScoringTable()
    fill_scoring_table.srcDir = sys.argv[1]
    fill_scoring_table.generateXlsx()
    fill_scoring_table.parserXlsx(fill_scoring_table.tablePath)
    fill_scoring_table.getFileName(fill_scoring_table.srcDir)
    wb = openpyxl.load_workbook(filename=fill_scoring_table.tablePath)
    sheets = wb.get_sheet_names()
    sheet1 = sheets[0]
    ws = wb.get_sheet_by_name(sheet1)
    for filename in fill_scoring_table.filename_to_row.keys():
        row_to_write = fill_scoring_table.filename_to_row[filename]
        for content_iter in range(2,len(fill_scoring_table.content_to_find)):
            col_to_write = content_iter + 3
            value_to_write = fill_scoring_table.adpater(filename,fill_scoring_table.content_to_find[content_iter])
            if value_to_write is None:
                continue
            ws.cell(row=row_to_write, column=col_to_write).value = value_to_write
    wb.save(filename=fill_scoring_table.tablePath)
